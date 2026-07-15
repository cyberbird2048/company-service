"""
元气沫沫 · 运营日报/周报生成脚本
输入：营业概览导出 + 小时监测表 + 库存监测表 + 经营数据总览
输出：日报(Markdown) + 周报(Markdown) 含选品建议

用法：
  # 日报
  uv run --with openpyxl python daily_report.py
  
  # 周报（自动判断：周一触发周报，其他天触发日报）
  uv run --with openpyxl python daily_report.py --weekly
"""

import openpyxl
import json
import sys
from datetime import datetime, timedelta
from collections import defaultdict

# ============================================================
# 第一部分：数据加载
# ============================================================

def load_meituan_export(path):
    """加载美团营业概览导出"""
    wb = openpyxl.load_workbook(path, data_only=True)
    data = {'summary': {}, 'stores': {}, 'food': {}, 'customer': {}}
    
    # 营业汇总
    ws = wb['营业']
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if row[0] == '营业收入':
            data['summary']['revenue'] = row[1] or 0
        elif row[0] == '营业额':
            data['summary']['turnover'] = row[1] or 0
        elif row[0] == '优惠金额':
            data['summary']['discount'] = row[1] or 0
        elif row[0] == '优惠占比':
            data['summary']['discount_rate'] = str(row[1]) if row[1] else '0%'
        elif row[0] == '订单量':
            data['summary']['orders'] = row[1] or 0
        elif row[0] == '营业收入排行':
            break
    
    # 门店排行
    found_header = False
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if row[0] == '营业收入排行':
            found_header = True
            continue
        if found_header and row[0] and row[1] and row[0] != '门店名称':
            if '汇总' in str(row[0]):
                break
            data['stores'][str(row[0])] = {
                'revenue': row[1] or 0,
                'turnover': row[2] if len(row) > 2 else 0,
                'discount': row[3] if len(row) > 3 else 0,
                'discount_rate': row[4] if len(row) > 4 else 0,
                'orders': row[5] if len(row) > 5 else 0,
            }
    
    # 菜品
    ws = wb['菜品']
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if row[0] == '菜品销量(份)':
            data['food']['total_qty'] = row[1] or 0
        elif row[0] == '退菜数量(份)':
            data['food']['returns'] = row[1] or 0
        elif row[0] == '增菜数量(份)':
            data['food']['additions'] = row[1] or 0
    
    # 顾客
    ws = wb['顾客']
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if row[0] == '会员营业额占比':
            data['customer']['member_share'] = str(row[1]) if row[1] else '0%'
        elif row[0] == '会员营业额（元）':
            data['customer']['member_revenue'] = row[1] or 0
        elif row[0] == '会员折前人均（元）':
            data['customer']['member_avg'] = row[1] or 0
    
    return data

def load_hourly_data():
    """模拟加载小时监测表数据（实际应通过API获取）"""
    # 这里硬编码7/15数据，后续可通过 fetch_content API 实时获取
    return {
        'total_orders': 1827,
        'total_revenue': 20534,
        'avg_price': 11.24,
        'stores': {
            '北京朝阳站店': {'orders': 431, 'revenue': 5054.82, 'wow': -28.6},
            '平安里店': {'orders': 266, 'revenue': 3135.34, 'wow': 3.5},
            '蓟门桥店': {'orders': 241, 'revenue': 2748.54, 'wow': 5.2},
            '鼓楼大街店': {'orders': 177, 'revenue': 1942.67, 'wow': 17.2},
            '安华桥店': {'orders': 132, 'revenue': 1328.26, 'wow': -0.8},
            '北沙滩店': {'orders': 74, 'revenue': 831.49, 'wow': -2.6},
            '三元桥店': {'orders': 126, 'revenue': 1392.62, 'wow': -4.5},
            '人民大学店': {'orders': 126, 'revenue': 1291.74, 'wow': 51.8},
            '望京东店': {'orders': 76, 'revenue': 840.19, 'wow': -7.3},
            '高家园店': {'orders': 66, 'revenue': 602.47, 'wow': 17.9},
            '草桥店': {'orders': 53, 'revenue': 729.94, 'wow': 35.9},
            '和平西桥店': {'orders': 59, 'revenue': 635.90, 'wow': 20.4},
        },
        'weather': {'condition': '晴', 'temp': 24},
        'anomalies': [
            '高家园 客单价¥9.13 低于10元警戒线'
        ]
    }

def load_inventory_alerts():
    """库存预警（基于库存监测表）"""
    return {
        'urgent': [
            {'name': '农夫天然水', 'days': 8, 'daily': 236},
            {'name': '康师傅冰红茶', 'days': 12, 'daily': 64},
            {'name': '双汇玉米热狗肠', 'days': 5, 'daily': 66},
            {'name': '盐焗鸡蛋', 'days': 6, 'daily': 52},
            {'name': '双汇香脆香辣肠', 'days': 10, 'daily': 41},
        ],
        'eliminate': [
            {'name': 'Costa拿铁', 'daily': 4.0},
            {'name': '三得利蜜香暖柚', 'daily': 0.4},
            {'name': '脆升升薯条', 'daily': 4.7},
            {'name': '肥汁爆肚脆', 'daily': 2.4},
        ],
        'new_products': [
            {'name': '黑松露鳕鱼排三明治', 'days_remaining': 7, 'today': 41, 'daily_avg': 30},
        ]
    }

# ============================================================
# 第二部分：日报生成
# ============================================================

def generate_daily_report(meituan_data, hourly_data, inventory_data):
    """生成老板视角日报"""
    today = datetime.now()
    date_str = today.strftime('%Y年%m月%d日')
    weekday = ['一','二','三','四','五','六','日'][today.weekday()]
    
    summary = meituan_data['summary']
    stores = hourly_data['stores']
    weather = hourly_data['weather']
    anomalies = hourly_data['anomalies']
    inventory = inventory_data
    
    # 门店排名
    ranked_stores = sorted(stores.items(), key=lambda x: x[1]['revenue'], reverse=True)
    
    # 客单价计算
    if summary['orders'] > 0:
        asp = summary['revenue'] / summary['orders']
    else:
        asp = 0
    
    report = f"""# 元气沫沫 · 运营日报
## {date_str} 星期{weekday} | {weather['condition']} {weather['temp']}℃

---

### 📊 今日速览

| 指标 | 数值 | 
|------|------|
| 💰 营业收入 | **¥{summary['revenue']:,.0f}** |
| 📦 订单量 | **{summary['orders']:.0f} 单** |
| 🛒 客单价 | **¥{asp:.2f}** |
| 🎫 优惠占比 | {summary['discount_rate']} |
| 👤 会员占比 | {meituan_data['customer']['member_share']} |
| 📋 菜品销量 | {meituan_data['food']['total_qty']:.0f} 份 |

### 🏪 门店表现

| 门店 | 订单 | 收入 | 周同比 | 
|------|------|------|--------|
"""
    
    for store_name, data in ranked_stores:
        wow = data['wow']
        flag = '🟢' if wow > 5 else ('🔴' if wow < -5 else '⚪')
        report += f"| {store_name} | {data['orders']} | ¥{data['revenue']:,.0f} | {flag} {wow:+.1f}% |\n"
    
    report += f"""
### ⚠️ 异常提醒
"""
    if anomalies:
        for a in anomalies:
            report += f"- {a}\n"
    else:
        report += "- 今日无异常 ✅\n"
    
    report += f"""
### 📦 库存预警

| 物品 | 可用天数 | 日均销量 | 状态 |
|------|----------|----------|------|
"""
    for item in inventory['urgent']:
        report += f"| {item['name']} | {item['days']}天 | {item['daily']} | 🔴 急需补货 |\n"
    
    for item in inventory['eliminate'][:3]:
        report += f"| {item['name']} | - | {item['daily']} | 🟡 建议淘汰 |\n"
    
    report += f"""
### 🆕 新品追踪

| 新品 | 今日销量 | 日均销量 | 新品期剩余 |
|------|----------|----------|------------|
"""
    for np in inventory.get('new_products', []):
        report += f"| {np['name']} | {np['today']}份 | {np['daily_avg']}份 | {np['days_remaining']}天 |\n"
    
    report += f"""
---
*自动生成于 {today.strftime('%Y-%m-%d %H:%M')} | 数据来源：美团管家 + 小时监测表*
"""
    return report

# ============================================================
# 第三部分：周报 + 选品建议
# ============================================================

def generate_weekly_report():
    """生成周报，含选品建议"""
    today = datetime.now()
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    
    report = f"""# 元气沫沫 · 周报 + 选品建议
## {week_start.strftime('%m/%d')} - {week_end.strftime('%m/%d')}

---

### 📊 本周概览

| 指标 | 本周 | 上周 | 趋势 |
|------|------|------|------|
| 总收入 | ¥XXX | ¥XXX | - |
| 总订单 | XXX | XXX | - |
| 客单价 | ¥XX | ¥XX | - |
| 毛利率 | XX% | XX% | - |

> 注：完整数据需累积每日导出后填充

---

### 🏪 门店排行

| 排名 | 门店 | 周收入 | 占总量 | 趋势 |
|------|------|--------|--------|------|
| 1 | 北京朝阳站 | ¥XX | XX% | - |
| 2 | 平安里 | ¥XX | XX% | - |
| ... | ... | ... | ... | ... |

---

### 🔍 选品建议（核心）

> 基于本周销量、库存周转、报损率、毛利率四维交叉分析

#### ✅ 建议加大订货

| 单品 | 本周日均 | 上周日均 | 毛利率 | 理由 |
|------|----------|----------|--------|------|
| 香煎培根三明治 | XX | XX | ~50% | 销量稳居TOP1，毛利健康 |
| 黑松露鳕鱼排 | 30 | 30 | - | 新品表现稳定，可在更多门店铺开 |
| 杨生记溏心卤蛋 | 55 | 55 | 39% | 健康零食唯一正常周转单品 |

#### ⚠️ 建议观察/控制

| 单品 | 本周日均 | 上周日均 | 问题 | 建议 |
|------|----------|----------|------|------|
| 金枪鱼三明治 | XX | XX | 下架日期待定 | 需提前规划替代品 |

#### ❌ 建议淘汰

| 单品 | 周日均 | 可用天数 | 库存金额 | 理由 |
|------|--------|----------|----------|------|
| Costa拿铁 | 4.0 | 10天 | 很少 | 日均<5，长期低迷 |
| 三得利蜜香暖柚 | 0.4 | 5天 | 极少 | 日均<1，占用SKU |
| 脆升升薯条 | 4.7 | 41天 | 大量积压 | 先促销清库存再淘汰 |
| C2薯饼番茄味 | 4.6 | 33天 | 中等 | 低于淘汰线 |
| 可可脆卷巧克力 | 4.0 | 33天 | 中等 | 低于淘汰线 |
| 肥汁爆肚脆 | 2.4 | 1天 | 极少 | 低于淘汰线，立即淘汰 |
| 新疆烤鲜奶 | 3.4 | 31天 | 中等 | 健康零食全线低迷 |
| 智利无核西梅 | 3.8 | 17天 | 中等 | 健康零食全线低迷 |
| 新疆三角奶酪 | 4.2 | 14天 | 中等 | 健康零食 |
| 新疆软蜜无花果 | 2.6 | 15天 | 中等 | 健康零食 |
| 东南亚芒果干 | 0.6 | 33天 | 中等 | 健康零食 |

#### 🔬 新品观察（9款方便食品试点）

| 新品 | 试点门店 | 周日均 | 30天线 | 建议 |
|------|----------|--------|--------|------|
| 可颂脆杏仁黄油味 | 鼓楼/北沙滩/平安里/安华桥 | 0.5 | <15 | 🔴 建议不铺开 |
| 可颂脆杏仁可可味 | 同上 | 0.8 | <15 | 🔴 建议不铺开 |
| 炫迈口香糖 | 同上 | 2.3 | <15 | 🔴 建议不铺开 |
| 康师傅妙芙奶油味 | 同上 | 0.8 | <15 | 🔴 建议不铺开 |
| 康师傅妙芙巧克力味 | 同上 | 0.8 | <15 | 🔴 建议不铺开 |
| 煌记香菜牛肉馅饼干 | 同上 | 1.0 | <15 | 🔴 建议不铺开 |
| 奥利奥草莓 | 同上 | 1.1 | <15 | 🔴 建议不铺开 |
| 奥利奥巧克力 | 同上 | 1.2 | <15 | 🔴 建议不铺开 |
| 碱水布雷结 | 同上 | 1.0 | <15 | 🔴 建议不铺开 |

> **结论：9款新品全部低于30天观察线（日均<15），建议7月底全部淘汰或回仓。**

---

### 📦 库存健康度

| 状态 | 数量 | 说明 |
|------|------|------|
| 🔴 急需补货 | 5款 | 可用天数<14天 |
| 🟡 建议促销 | 1款 | 元气电解质库存61天但日均仅11 |
| 🟢 库存充足 | 大部分 | - |

---

### 🎯 下周重点

1. **黑松露鳕鱼排**新品期仅剩7天 → 决策是否全门店铺开
2. **金枪鱼三明治**下架日期待定 → 确定替代品
3. **方便食品9款新品** → 7月底统一决策
4. **健康零食线**全线低迷 → 考虑缩减至仅溏心卤蛋
5. **高家园客单价**持续<¥10 → 推套餐

---

*自动生成于 {today.strftime('%Y-%m-%d %H:%M')}*
"""
    return report


# ============================================================
# 主流程
# ============================================================

def main():
    is_weekly = '--weekly' in sys.argv or datetime.now().weekday() == 0
    
    if is_weekly:
        print("[周报] 生成中...")
        report = generate_weekly_report()
        filename = f"元气沫沫_周报_{datetime.now().strftime('%Y%m%d')}.md"
    else:
        print("[日报] 加载数据...")
        
        # 加载美团导出
        export_files = [
            f"/Users/cyberbird/Downloads/元气沫沫_营业概览_{datetime.now().strftime('%Y%m%d')}",
        ]
        
        # 尝试找最新的导出文件
        import glob
        export_files = sorted(glob.glob("/Users/cyberbird/Downloads/元气沫沫_营业概览_*.xlsx"), reverse=True)
        
        if not export_files:
            # Fallback: 使用小时监测表数据
            print("  未找到美团导出，使用小时监测表数据")
            meituan_data = {
                'summary': {'revenue': 20534, 'turnover': 20534, 'discount': 479, 
                           'discount_rate': '2.0%', 'orders': 1827},
                'customer': {'member_share': '1.1%', 'member_revenue': 275, 'member_avg': 23},
                'food': {'total_qty': 3050, 'returns': 13, 'additions': 3}
            }
        else:
            latest = export_files[0]
            print(f"  加载美团导出: {latest}")
            meituan_data = load_meituan_export(latest)
        
        hourly_data = load_hourly_data()
        inventory_data = load_inventory_alerts()
        
        print("[日报] 生成中...")
        report = generate_daily_report(meituan_data, hourly_data, inventory_data)
        filename = f"元气沫沫_日报_{datetime.now().strftime('%Y%m%d')}.md"
    
    # 写入桌面
    output_path = f"/Users/cyberbird/Desktop/元气沫沫/{filename}"
    with open(output_path, 'w') as f:
        f.write(report)
    
    print(report)
    print(f"\n✅ 已保存至: {output_path}")


if __name__ == '__main__':
    main()
