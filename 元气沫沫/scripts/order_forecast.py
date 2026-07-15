"""
元气沫沫 · 智能订货脚本 v0.1
输入：经营数据总览.xlsx + 库存监测表 + 天气数据
输出：每门店每品类每日订货建议

用法：
  codex exec python order_forecast.py
  或 python3 order_forecast.py
"""

import openpyxl
import json
import sys
from datetime import datetime, timedelta
from collections import defaultdict

# ============================================================
# 第一部分：数据加载
# ============================================================

def load_operational_data(path):
    """加载经营数据总览，返回 per-store per-month 的数据"""
    wb = openpyxl.load_workbook(path, data_only=True)
    stores = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == '总计':
            continue
        ws = wb[sheet_name]
        months = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            month, orders, qty, rev, discount, disc_rate = row[0], row[1], row[2], row[3], row[4], row[5]
            if month and orders and orders > 0:
                months.append({
                    'month': str(month),
                    'orders': orders,
                    'quantity': qty,
                    'revenue': rev,
                    'discount': discount,
                    'discount_rate': disc_rate,
                })
        if months:
            stores[sheet_name] = months
    return stores

def load_inventory_data(path):
    """加载库存监测表数据"""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        items = []
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            if row[0] and row[1]:
                items.append({
                    'name': row[0],
                    'category': row[1],
                    'warehouse_stock': row[6] or 0,
                    'store_stock': row[7] or 0,
                    'daily_avg_sales': row[8] or 0,
                    'available_days': row[9] or 0,
                    'suggestion': row[10] if len(row) > 10 else '',
                })
        return items
    except Exception as e:
        print(f"[WARN] 库存数据加载失败: {e}")
        return []

# ============================================================
# 第二部分：核心算法
# ============================================================

def forecast_daily_sales(store_data, category, lookback_months=3):
    """
    基于历史数据预测日均销量
    策略：取最近N个月的移动平均，按品类系数折算
    """
    months = store_data
    if len(months) < lookback_months:
        lookback_months = len(months)
    
    recent = months[-lookback_months:]
    total_orders = sum(m['orders'] for m in recent)
    total_days = sum(30 for m in recent)  # 简化：每月30天
    daily_avg = total_orders / total_days if total_days > 0 else 0
    
    # 品类系数（基于6月运营报告中各品类收入占比+毛利率）
    category_ratios = {
        '三明治': 0.224,      # 22.4% 收入占比
        '热狗汉堡': 0.112,     # 11.2%
        '手作面包': 0.173,     # 17.3%
        '手工吐司': 0.087,     # 8.7%
        '缤纷饮品': 0.213,     # 21.3%
        '方便食品': 0.105,     # 10.5%
    }
    
    return daily_avg * category_ratios.get(category, 0.10)

def apply_weather_factor(base_forecast, weather):
    """天气因子修正"""
    factor = 1.0
    temp = weather.get('temp', 25)
    is_rain = weather.get('rain', False)
    is_weekend = weather.get('is_weekend', False)
    
    # 高温 → 饮品需求上升
    if temp > 35:
        if weather.get('category') == '缤纷饮品':
            factor *= 1.3
    # 雨天工作日 → 进店减少
    if is_rain and not is_weekend:
        factor *= 0.85
    # 雨天周末 → 出行大幅减少
    if is_rain and is_weekend:
        factor *= 0.5
    # 周末 → 仅限全周营业门店
    if is_weekend:
        if weather.get('store_opens_weekend', True):
            factor *= 0.7  # 周末客流约为工作日70%
        else:
            factor = 0
    
    return base_forecast * factor

def apply_holiday_factor(base_forecast, holiday_info):
    """节假日因子修正"""
    phase = holiday_info.get('phase', 'normal')
    factors = {
        'pre_holiday_2d': 1.5,   # 节前2天
        'pre_holiday_1d': 1.3,   # 节前1天
        'holiday': 1.2,          # 节日当天
        'post_holiday_1d': 0.7,  # 节后1天
        'post_holiday_2d': 0.85, # 节后2天
        'normal': 1.0,
    }
    return base_forecast * factors.get(phase, 1.0)

def apply_loss_control(order_suggestion, store_loss_rate):
    """报损率约束：报损率>5%的门店自动减量20%"""
    if store_loss_rate > 0.05:
        return order_suggestion * 0.80
    return order_suggestion

# ============================================================
# 第三部分：输出生成
# ============================================================

def generate_order_sheet(stores_data, inventory_items, weather, holiday_info, store_loss_rates):
    """生成订货建议表"""
    
    # 品类→跨店平均销量缓存
    category_avg = defaultdict(float)
    
    results = []
    for store_name, store_months in stores_data.items():
        if len(store_months) < 2:
            continue
        active_stores = ["北京朝阳站店","平安里店","蓟门桥店","鼓楼大街店","安华桥店","北沙滩店","三元桥店","人民大学店","望京东店","高家园店","草桥店","和平西桥店"]
        if store_name not in active_stores:
            continue

            continue
        
        store_result = {
            'store': store_name,
            'categories': {},
            'summary': {}
        }
        
        for category in ['三明治', '热狗汉堡', '手作面包', '手工吐司']:
            base = forecast_daily_sales(store_months, category)
            
            # 合并天气和节假日因子
            weather['category'] = category
            weather['store_opens_weekend'] = store_name not in [
                '三元桥店', '人民大学店', '草桥店', '望京东店', '高家园店', '和平西桥店'
            ]
            adjusted = apply_weather_factor(base, weather)
            adjusted = apply_holiday_factor(adjusted, holiday_info)
            
            # 报损控制
            loss_rate = store_loss_rates.get(store_name, 0.03)
            final = apply_loss_control(adjusted, loss_rate)
            
            # 向上取整
            final_qty = max(1, round(final)) if final > 0 else 0
            
            store_result['categories'][category] = {
                'base_forecast': round(base, 1),
                'adjusted_forecast': round(adjusted, 1),
                'suggested_order': final_qty,
                'loss_rate': f"{loss_rate*100:.1f}%",
            }
        
        total_order = sum(c['suggested_order'] for c in store_result['categories'].values())
        store_result['summary'] = {
            'total_daily_order': total_order,
            'avg_daily_orders': round(sum(m['orders'] for m in store_months[-3:]) / 90, 1),
        }
        
        results.append(store_result)
    
    return results

def print_order_sheet(results, holiday_info, weather):
    """格式化输出订货建议表"""
    today = datetime.now().strftime('%Y-%m-%d')
    holiday_phase = holiday_info.get('phase', 'normal')
    temp = weather.get('temp', 'N/A')
    rain = '雨' if weather.get('rain') else '晴'
    
    print(f"=" * 80)
    print(f"  元气沫沫 · 智能订货建议表")
    print(f"  生成时间: {today}")
    print(f"  天气: {rain} {temp}°C | 节假日: {holiday_phase}")
    print(f"=" * 80)
    
    # 库存预警
    print(f"\n⚠️  库存补货预警:")
    print(f"  {'物品':<20} {'可用天数':>6} {'日均销量':>8} {'状态'}")
    print(f"  {'-'*50}")
    
    # 简化：对关键品类做检查
    key_items = {
        '双汇玉米热狗肠': '热狗汉堡',
        '热狗肠多拿滋': '热狗汉堡',
        '培根帕尼尼(大)': '热狗汉堡',
    }
    for item_name, category in key_items.items():
        print(f"  {item_name:<20} {'--':>6} {'--':>8} 需从库存表获取")
    
    print(f"\n📋 各门店订货建议:")
    print(f"  {'门店':<12} {'三明治':>6} {'热狗汉堡':>6} {'手作面包':>6} {'手工吐司':>6} {'合计':>6} {'报损率'}")
    print(f"  {'-'*65}")
    
    for r in results:
        cats = r['categories']
        items = []
        for c in ['三明治', '热狗汉堡', '手作面包', '手工吐司']:
            items.append(str(cats[c]['suggested_order']) if cats[c]['suggested_order'] > 0 else "-")
        loss = cats['三明治']['loss_rate']
        total = r['summary']['total_daily_order']
        print(f"  {r['store']:<12} {items[0]:>6} {items[1]:>6} {items[2]:>6} {items[3]:>6} {total:>6} {loss:>7}")
    
    print(f"\n📝 备注:")
    print(f"  1. 以上为建议订货量，请店长根据实际陈列空间微调")
    print(f"  2. 短保品（三明治/面包）按 T+2 到货节奏，订货量为单日需求")
    print(f"  3. 朝阳站节假日备货上限不超过平日2倍")
    if weather.get('rain') and not weather.get('is_weekend'):
        print(f"  4. 今日有雨，工作日进店减少15%，已自动下调")

# ============================================================
# 第四部分：主流程
# ============================================================

def main():
    # 配置路径
    DATA_FILE = "/Users/cyberbird/Desktop/元气沫沫/元气沫沫经营数据总览.xlsx"
    
    # 加载数据
    print("[1/4] 加载经营数据...")
    stores = load_operational_data(DATA_FILE)
    print(f"      已加载 {len(stores)} 家门店")
    
    # 今日天气（可通过API获取，此处硬编码示例）
    # 实际可从小时监测表页面获取天气数据
    weather = {
        'temp': 24,
        'rain': False,
        'is_weekend': datetime.now().weekday() >= 5,
    }
    
    # 节假日检查
    holiday_info = {'phase': 'normal'}
    
    # 各门店报损率（来自6月运营报告）
    store_loss_rates = {
        '北京朝阳站店': 0.022,
        '平安里店': 0.043,
        '蓟门桥店': 0.024,
        '鼓楼大街店': 0.061,
        '安华桥店': 0.031,
        '北沙滩店': 0.035,
        '三元桥店': 0.027,
        '人民大学店': 0.033,
        '望京东店': 0.033,
        '高家园店': 0.082,
        '草桥店': 0.081,
        '和平西桥店': 0.023,
    }
    
    # 生成订货建议
    print("[2/4] 执行需求预测...")
    results = generate_order_sheet(stores, [], weather, holiday_info, store_loss_rates)
    
    # 输出结果
    print("[3/4] 生成订货表...")
    print_order_sheet(results, holiday_info, weather)
    
    # 输出JSON（供后续程序消费）
    output = {
        'generated_at': datetime.now().isoformat(),
        'weather': weather,
        'holiday': holiday_info,
        'orders': [
            {
                'store': r['store'],
                'categories': r['categories'],
                'summary': r['summary'],
            }
            for r in results
        ]
    }
    
    output_path = "/Users/cyberbird/Desktop/元气沫沫/订货建议_%s.json" % datetime.now().strftime('%Y%m%d')
    with open(output_path, 'w') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f"\n[4/4] 结果已保存至: {output_path}")

if __name__ == '__main__':
    main()
